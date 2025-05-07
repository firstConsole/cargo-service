import re
import sys
import os
import logging
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from fuzzywuzzy import fuzz
from tabulate import tabulate
from typing import Dict, List, Any, Optional
from deep_translator import GoogleTranslator

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)
logger = logging.getLogger('excel_parser')


class CargoExcelParser:
    """
    Улучшенный парсер Excel файлов для данных о грузах.
    Использует pandas, fuzzywuzzy и другие библиотеки для более эффективного анализа.
    """

    def __init__(self, file_path: str, period_id: str = "unknown"):
        self.file_path = file_path
        self.period_id = period_id
        self.workbook = None
        self.sheet = None
        self.df = None
        self.batch_number = None
        self.batch_number_numeric = None
        self.data_start_row = None
        self.data_end_row = None
        self.column_mapping = {}
        self.parsed_data = []
        self.batch_summary = {}  # Добавляем поле для хранения сводной информации

        # Словарь для нечеткого сопоставления заголовков столбцов
        self.column_patterns = {
            'clientCode': ['Литер', 'литер', 'Литер客户代码/日期', '客户代码/日期'],
            'depatureFromChinaDate': ['客户代码/日期'],
            'placesCount': ['мест', 'Мест', 'Мест包数', '包数'],
            'weight': ['вес', 'Вес', ' Вес重量', 'Вес重量', '重量'],
            'boxesCount': ['коробки', 'кор', 'Коробки', 'Кор', 'Кор箱数', '箱数'],
            'cubicTariff': ['Цена运价', 'Цена', '运价'],
            'unitsCount': ['шт', 'Шт', '件数', 'Шт件数'],
            'productName': ['наименование', '品名', 'наименование品名'],
            'productPrice': ['1ед', '货值', 'цена', ' 1ед货值', '1ед货值'],
            'insurancePercent': ['%', '%保率', '保率'],
            'volume': ['Куб', 'куб', '体积', 'Куб体积'],
            'freightTariff': ['фрахт', 'Фрахт', '运费', 'фрахт运费', 'Фрахт运费'],
            'insurance': ['Стр-ка', 'стр-ка', 'страховка', 'Страховка', '保险费', 'Стр-ка保险费'],
            'packaging': ['Упа-ка', 'упа-ка', 'упаковка', 'Упаковка', '包装费', 'Упа-ка包装费', 'упа-ка包装费'],
            'total': ['итого', 'Итого', '总计$', '$', 'Итого总计$']
        }

    def parse(self) -> List[Dict[str, Any]]:
        """
        Основной метод парсинга файла Excel.

        Returns:
            List[Dict[str, Any]]: Список словарей с данными о грузах
        """
        try:
            logger.info(f"Начинаем парсинг файла: {self.file_path}")

            # Пробуем сначала загрузить через pandas для предварительного анализа
            try:
                self.df = pd.read_excel(self.file_path, header=None)
                logger.info(f"Файл успешно загружен через pandas. Размер: {self.df.shape}")
            except Exception as e:
                logger.warning(f"Не удалось загрузить файл через pandas: {e}. Используем openpyxl.")
                self.df = None

            # Загружаем Excel файл через openpyxl для доступа к объединенным ячейкам
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            self.sheet = self.workbook.active

            # Находим номер партии (баланса)
            self._find_batch_number()
            if not self.batch_number:
                logger.error("Номер партии (баланса) не найден. Парсинг прерван.")
                return []

            # Находим диапазон данных
            self._find_data_range()
            if not self.data_start_row:
                logger.error("Не удалось определить начало данных. Парсинг прерван.")
                return []

            # Определяем соответствие столбцов
            self._map_columns()

            # Парсим данные строк
            self._parse_data_rows()

            # Рассчитываем сводную информацию по партии
            self.batch_summary = self._calculate_batch_summary()

            logger.info(f"Парсинг завершен. Найдено {len(self.parsed_data)} записей.")

            # Выводим результаты в консоль
            self._print_results()

            return self.parsed_data

        except Exception as e:
            logger.error(f"Ошибка при парсинге файла: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return []

    def _find_batch_number(self) -> None:
        """
        Находит номер партии (баланса) в Excel файле.
        Использует как точное, так и нечеткое сопоставление.
        """
        # Если есть DataFrame, используем его для поиска
        if self.df is not None:
            # Ищем в первых 10 строках и 10 столбцах
            search_area = self.df.iloc[:10, :10]

            # Преобразуем все в строки для поиска
            search_area = search_area = search_area.astype(str)

            # Ищем по шаблону
            for i in range(search_area.shape[0]):
                for j in range(search_area.shape[1]):
                    value = search_area.iloc[i, j]
                    batch_number = self._extract_batch_number(value)
                    if batch_number:
                        self.batch_number = batch_number
                        self.batch_number_numeric = self._extract_numeric_part(batch_number)
                        logger.info(
                            f"Найден номер партии в ячейке [{i}, {j}]: {batch_number} (числовая часть: {self.batch_number_numeric})")
                        return

        # Если не нашли через pandas или его нет, используем openpyxl
        standard_cell = self.sheet['D2']
        batch_number = self._extract_batch_number(standard_cell.value)

        if batch_number:
            self.batch_number = batch_number
            self.batch_number_numeric = self._extract_numeric_part(batch_number)
            logger.info(
                f"Найден номер партии в стандартном месте: {batch_number} (числовая часть: {self.batch_number_numeric})")
            return

        # Если не нашли в стандартном месте, ищем в первых 10 строках столбцов от A до J
        for col in range(1, 11):
            col_letter = get_column_letter(col)
            for row in range(1, 11):
                cell = self.sheet[f"{col_letter}{row}"]
                batch_number = self._extract_batch_number(cell.value)
                if batch_number:
                    self.batch_number = batch_number
                    self.batch_number_numeric = self._extract_numeric_part(batch_number)
                    logger.info(
                        f"Найден номер партии в ячейке {col_letter}{row}: {batch_number} (числовая часть: {self.batch_number_numeric})")
                    return

        # Если не нашли по точному шаблону, пробуем нечеткий поиск
        batch_patterns = ['баланс', 'партия', '单据名称：', 'balance']

        for col in range(1, 11):
            col_letter = get_column_letter(col)
            for row in range(1, 11):
                cell_value = str(self.sheet[f"{col_letter}{row}"].value or "")

                # Проверяем, содержит ли ячейка ключевые слова
                if any(pattern in cell_value.lower() for pattern in batch_patterns):
                    # Ищем в этой и соседних ячейках
                    for r in range(max(1, row - 1), min(row + 2, 11)):
                        for c in range(max(1, col - 1), min(col + 2, 11)):
                            c_letter = get_column_letter(c)
                            neighbor_value = str(self.sheet[f"{c_letter}{r}"].value or "")

                            # Проверяем на соответствие шаблону
                            batch_number = self._extract_batch_number(neighbor_value)
                            if batch_number:
                                self.batch_number = batch_number
                                self.batch_number_numeric = self._extract_numeric_part(batch_number)
                                logger.info(
                                    f"Найден номер партии через нечеткий поиск в ячейке {c_letter}{r}: {batch_number}")
                                return

        logger.warning("Номер партии не найден")

    @staticmethod
    def _extract_batch_number(value) -> Optional[str]:
        """
        Извлекает номер партии из значения ячейки.

        Args:
            value: Значение ячейки

        Returns:
            Optional[str]: Номер партии или None
        """
        if not value:
            return None

        value = str(value).strip()

        # Пробуем разные шаблоны
        patterns = [
            r'([A-Za-z][0-9]{1,5}-[A-Za-z])',  # Основной шаблон
            r'([A-Za-z][0-9]{1,5}/[A-Za-z])',  # Вариант с /
            r'([A-Za-z][0-9]{1,5}\s+[A-Za-z])'  # Вариант с пробелом
        ]

        for pattern in patterns:
            match = re.search(pattern, value)
            if match:
                return match.group(1)

        return None

    @staticmethod
    def _extract_numeric_part(batch_number: str) -> str:
        """
        Извлекает числовую часть из номера партии.

        Args:
            batch_number: Номер партии

        Returns:
            str: Числовая часть номера партии
        """
        match = re.search(r'[A-Za-z]([0-9]{1,5})[-/\s][A-Za-z]', batch_number)
        if match:
            return match.group(1)
        return ""

    @staticmethod
    def _extract_client_code_numeric(client_code: str) -> str:
        """
        Извлекает числовую часть из кода клиента.

        Args:
            client_code: Код клиента

        Returns:
            str: Числовая часть кода клиента
        """
        if not client_code:
            return ""

        # Извлекаем все цифры из кода клиента
        match = re.search(r'(\d+)', client_code)
        if match:
            return match.group(1)
        return ""

    def _find_data_range(self) -> None:
        """
        Находит диапазон строк с данными в Excel файле.
        Использует эвристики для определения начала и конца данных.
        """
        # Если есть DataFrame, используем его для поиска
        if self.df is not None:
            # Ищем строки, где в первом столбце есть код клиента
            client_code_pattern = r'^[A-Z]{1,4}[0-9]{1,7}[A-Z]{0,2}$'

            # Преобразуем первый столбец в строки
            first_col = self.df.iloc[:, 0].astype(str)

            # Находим строки, соответствующие шаблону
            matches = first_col.str.match(client_code_pattern)

            if matches.any():
                # Находим индексы строк с данными
                data_rows = matches[matches].index.tolist()

                if data_rows:
                    # Проверяем, что строки идут последовательно
                    consecutive_groups = []
                    current_group = [data_rows[0]]

                    for i in range(1, len(data_rows)):
                        if data_rows[i] == data_rows[i - 1] + 1:
                            current_group.append(data_rows[i])
                        else:
                            consecutive_groups.append(current_group)
                            current_group = [data_rows[i]]

                    consecutive_groups.append(current_group)

                    # Берем самую длинную последовательность
                    longest_group = max(consecutive_groups, key=len)

                    self.data_start_row = longest_group[0] + 1
                    self.data_end_row = longest_group[-1] + 1

                    logger.info(
                        f"Найден диапазон данных через pandas: строки {self.data_start_row}-{self.data_end_row}")
                    return

        # Если не нашли через pandas или его нет, используем openpyxl
        client_code_pattern = r'^[A-Z]{1,4}[0-9]{1,7}[A-Z]{0,2}$'

        start_row = None
        end_row = None

        # Ищем последовательные строки с кодами клиентов
        consecutive_rows = []
        current_sequence = []

        for row in range(1, self.sheet.max_row + 1):
            cell_value = self.sheet.cell(row=row, column=1).value
            if cell_value and isinstance(cell_value, str) and re.match(client_code_pattern, cell_value.strip()):
                current_sequence.append(row)
            elif current_sequence:
                consecutive_rows.append(current_sequence)
                current_sequence = []

        if current_sequence:
            consecutive_rows.append(current_sequence)

        # Выбираем самую длинную последовательность
        if consecutive_rows:
            longest_sequence = max(consecutive_rows, key=len)
            if len(longest_sequence) > 1:
                start_row = longest_sequence[0]
                end_row = longest_sequence[-1]
                logger.info(f"Найден диапазон данных: строки {start_row}-{end_row}")

        # Если не нашли последовательность, ищем по одиночным совпадениям
        if not start_row:
            for row in range(1, self.sheet.max_row + 1):
                cell_value = self.sheet.cell(row=row, column=1).value
                if cell_value and isinstance(cell_value, str) and re.match(client_code_pattern, cell_value.strip()):
                    if start_row is None:
                        start_row = row
                        logger.info(f"Найдено начало данных в строке {row}")
                elif start_row is not None and not end_row:
                    # Проверяем следующие строки, чтобы убедиться, что это действительно конец данных
                    next_rows_empty = True
                    for check_row in range(row, min(row + 3, self.sheet.max_row + 1)):
                        check_value = self.sheet.cell(row=check_row, column=1).value
                        if check_value and isinstance(check_value, str) and re.match(client_code_pattern,
                                                                                     check_value.strip()):
                            next_rows_empty = False
                            break

                    if next_rows_empty:
                        end_row = row - 1
                        logger.info(f"Найден конец данных в строке {end_row}")
                        break

        # Если нашли начало, но не нашли конец, считаем концом последнюю строку
        if start_row and not end_row:
            end_row = self.sheet.max_row
            logger.info(f"Конец данных не найден явно, используем последнюю строку: {end_row}")

        self.data_start_row = start_row
        self.data_end_row = end_row

    def _map_columns(self) -> None:
        """
        Определяет соответствие столбцов в Excel файле.
        Использует как точное, так и нечеткое сопоставление.
        """
        if not self.data_start_row:
            return

        # Начинаем с известных столбцов
        self.column_mapping = {
            'clientCode': 1  # Первый столбец всегда код клиента
        }

        # Если есть заголовки, пробуем их использовать
        header_row = self.data_start_row - 1 if self.data_start_row > 1 else None

        if header_row:
            # Собираем заголовки
            headers = {}
            for col in range(1, self.sheet.max_column + 1):
                header_value = self.sheet.cell(row=header_row, column=col).value
                if header_value:
                    headers[col] = str(header_value).lower().strip()

            # Сопоставляем заголовки с нашими полями
            for field, patterns in self.column_patterns.items():
                best_match = None
                best_score = 0

                for col, header in headers.items():
                    # Проверяем точное совпадение
                    if any(pattern in header for pattern in patterns):
                        best_match = col
                        best_score = 100
                        break

                    # Есл�� нет точного совпадения, используем нечеткое
                    for pattern in patterns:
                        score = fuzz.partial_ratio(pattern, header)
                        if score > best_score:
                            best_score = score
                            best_match = col

                # Если нашли совпадение с высоким рейтингом
                if best_match and best_score > 70:
                    self.column_mapping[field] = best_match
                    logger.info(
                        f"Сопоставлено поле {field} со столбцом {best_match} (заголовок: {headers.get(best_match)}, рейтинг: {best_score})")

        # Если не удалось сопоставить все поля через заголовки, используем эвристики

        # Ищем столбец с датой выхода из Китая (обычно столбец C)
        # Принудительно устанавливаем столбец C (индекс 3) для даты выхода из Китая
        departure_col = 3
        self.column_mapping['depatureFromChinaDate'] = departure_col

        # Последовательно определяем остальные столбцы, если они еще не определены
        if 'placesCount' not in self.column_mapping:
            self.column_mapping['placesCount'] = departure_col + 1

        if 'weight' not in self.column_mapping:
            self.column_mapping['weight'] = departure_col + 2

        if 'boxesCount' not in self.column_mapping:
            self.column_mapping['boxesCount'] = departure_col + 3

        if 'cubicTariff' not in self.column_mapping:
            self.column_mapping['cubicTariff'] = departure_col + 4

        if 'unitsCount' not in self.column_mapping:
            self.column_mapping['unitsCount'] = departure_col + 5

        if 'productPrice' not in self.column_mapping:
            self.column_mapping['productPrice'] = departure_col + 6

        if 'insurancePercent' not in self.column_mapping:
            self.column_mapping['insurancePercent'] = departure_col + 7

        if 'volume' not in self.column_mapping:
            self.column_mapping['volume'] = departure_col + 8

        if 'productName' not in self.column_mapping:
            self.column_mapping['productName'] = departure_col + 9

        current_col = max(self.column_mapping.values()) + 1

        # Ищем freightTariff (формат "182$")
        if 'freightTariff' not in self.column_mapping:
            freight_col = self._find_column_by_format(
                start_col=current_col,
                pattern=r'^\d+(\.\d+)?\$$',
                sample_row=self.data_start_row
            )

            if freight_col:
                self.column_mapping['freightTariff'] = freight_col
                current_col = freight_col + 1
        else:
            current_col = self.column_mapping['freightTariff'] + 1

        # Ищем insurance (формат "7$")
        if 'insurance' not in self.column_mapping:
            self.column_mapping['insurance'] = current_col
            current_col += 1

        # Ищем packaging (формат "51$")
        if 'packaging' not in self.column_mapping:
            self.column_mapping['packaging'] = current_col
            current_col += 1

        # Последний столбец - total
        if 'total' not in self.column_mapping:
            # Ищем последний столбец с данными
            last_col = self._find_last_data_column()
            self.column_mapping['total'] = last_col

        logger.info(f"Определено соответствие столбцов: {self.column_mapping}")

    def _find_column_by_format(self, start_col: int, pattern: str, sample_row: int) -> Optional[int]:
        """
        Находит столбец, значения в котором соответствуют заданному шаблону.

        Args:
            start_col: Начальный столбец для поиска
            pattern: Регулярное выражение для поиска
            sample_row: Строка для проверки

        Returns:
            Optional[int]: Номер столбца или None
        """
        max_col = self.sheet.max_column

        for col in range(start_col, max_col + 1):
            cell_value = self.sheet.cell(row=sample_row, column=col).value
            if cell_value and isinstance(cell_value, str):
                if re.match(pattern, cell_value.strip()):
                    return col

        return None

    def _find_last_data_column(self) -> int:
        """
        Находит последний столбец с данными.

        Returns:
            int: Номер последнего столбца с данными
        """
        max_col = self.sheet.max_column

        # Проверяем наличие данных в строках
        for col in range(max_col, 0, -1):
            for row in range(self.data_start_row, self.data_end_row + 1):
                cell_value = self.sheet.cell(row=row, column=col).value
                if cell_value:
                    return col

        # Если не нашли, возвращаем последний столбец
        return max_col

    def _parse_data_rows(self) -> None:
        """
        Парсит данные из строк Excel файла.
        Обрабатывает объединенные ячейки и сборные места.
        """
        if not self.data_start_row or not self.data_end_row:
            return

        # Определяем объединенные ячейки для обработки сборных мест
        merged_ranges = self.sheet.merged_cells.ranges
        merged_cells_map = {}
        composite_cargo_groups = {}  # Словарь для группировки строк сборных мест

        # Создаем карту объединенных ячеек
        for merged_range in merged_ranges:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_cells_map[(row, col)] = {
                        'range': merged_range,
                        'main_cell': (merged_range.min_row, merged_range.min_col)
                    }

        # Первый проход: определяем сборные места и их группы
        places_col = self.column_mapping.get('placesCount')
        if places_col:
            for row in range(self.data_start_row, self.data_end_row + 1):
                cell_key = (row, places_col)
                if cell_key in merged_cells_map:
                    merged_info = merged_cells_map[cell_key]
                    merged_range = merged_info['range']
                    main_cell_key = merged_info['main_cell']

                    # Создаем уникальный идентификатор для группы сборного места
                    group_id = f"composite_{main_cell_key[0]}_{main_cell_key[1]}"

                    # Если это основная ячейка, создаем новую группу
                    if cell_key == main_cell_key:
                        composite_cargo_groups[group_id] = {
                            'main_row': row,
                            'rows': [],
                            'places_count': merged_range.max_row - merged_range.min_row + 1
                        }

                    # Добавляем текущую строку в группу
                    if group_id in composite_cargo_groups:
                        composite_cargo_groups[group_id]['rows'].append(row)

        # Добавляем счетчик для порядковых номеров сборников
        composite_counter = 1

        # Для каждой группы сборных мест добавим уникальный ID
        for group_id, group_info in composite_cargo_groups.items():
            places_count = group_info['places_count']
            composite_id = f"{self.period_id}-{self.batch_number_numeric}-{places_count}-{composite_counter}"
            group_info['composite_id'] = composite_id
            composite_counter += 1

        # Второй проход: парсим данные из всех строк
        for row in range(self.data_start_row, self.data_end_row + 1):
            # Проверяем, является ли эта строка частью сборного места
            is_composite = False
            composite_group_id = None
            composite_places_count = 1
            is_main_row = False
            composite_id = None

            # Проверяем все группы сборных мест
            for group_id, group_info in composite_cargo_groups.items():
                if row in group_info['rows']:
                    is_composite = True
                    composite_group_id = group_id
                    composite_places_count = group_info['places_count']
                    is_main_row = (row == group_info['main_row'])
                    composite_id = group_info.get('composite_id')
                    break

            # Инициализируем данные строки
            row_data = {
                'batchNumber': self.batch_number,
                'batchNumberNumeric': self.batch_number_numeric,
                'isCompositeCargo': is_composite,
                'isMainCompositeRow': is_main_row,
                'compositeGroupId': composite_group_id,
                'compositePlacesCount': composite_places_count if is_composite else 1,
                'compositeId': composite_id
            }

            # Парсим данные по столбцам
            for field, col in self.column_mapping.items():
                cell = self.sheet.cell(row=row, column=col)

                # Если это объединенная ячейка, берем значение из основной ячейки
                if isinstance(cell, MergedCell) or (row, col) in merged_cells_map:
                    if (row, col) in merged_cells_map:
                        main_row, main_col = merged_cells_map[(row, col)]['main_cell']
                        cell = self.sheet.cell(row=main_row, column=main_col)

                value = cell.value

                # Обработка значений в зависимости от поля
                if field == 'clientCode':
                    if value and isinstance(value, str):
                        row_data[field] = value.strip()
                        # Добавляем числовую часть кода клиента
                        row_data['clientCodeNumeric'] = self._extract_client_code_numeric(value.strip())
                    else:
                        row_data[field] = None
                        row_data['clientCodeNumeric'] = None

                elif field == 'depatureFromChinaDate':
                    if value and isinstance(value, str):
                        # Проверяем формат "0123" (4 цифры)
                        if re.match(r'^\d{4}$', value.strip()):
                            row_data[field] = value.strip()
                        else:
                            row_data[field] = value.strip()
                    else:
                        row_data[field] = value

                elif field == 'placesCount':
                    if value is not None:
                        try:
                            row_data[field] = int(value)
                        except (ValueError, TypeError):
                            row_data[field] = value
                    else:
                        row_data[field] = None

                elif field == 'weight':
                    if value is not None:
                        try:
                            # Преобразуем в число с плавающей запятой
                            if isinstance(value, str):
                                # Заменяем запятую на точку для корректного преобразования
                                value = value.replace(',', '.')
                            row_data[field] = float(value)
                        except (ValueError, TypeError):
                            row_data[field] = value
                    else:
                        row_data[field] = None

                elif field == 'boxesCount':
                    if value is not None:
                        try:
                            row_data[field] = int(value)
                        except (ValueError, TypeError):
                            row_data[field] = value
                    else:
                        row_data[field] = None

                elif field == 'cubicTariff':
                    if value and isinstance(value, str):
                        match = re.search(r'(\d+(\.\d+)?)', value)
                        if match:
                            row_data[field] = float(match.group(1))
                        else:
                            row_data[field] = value
                    elif value is not None:
                        try:
                            row_data[field] = float(value)
                        except (ValueError, TypeError):
                            row_data[field] = value
                    else:
                        row_data[field] = None

                elif field == 'unitsCount':
                    if value is not None:
                        try:
                            row_data[field] = int(value)
                        except (ValueError, TypeError):
                            row_data[field] = value
                    else:
                        row_data[field] = None

                elif field == 'productPrice':
                    if value and isinstance(value, str):
                        match = re.search(r'(\d+(\.\d+)?)', value)
                        if match:
                            row_data[field] = float(match.group(1))
                        else:
                            row_data[field] = value
                    elif value is not None:
                        try:
                            row_data[field] = float(value)
                        except (ValueError, TypeError):
                            row_data[field] = value
                    else:
                        row_data[field] = None

                elif field == 'insurancePercent':
                    if value and isinstance(value, str):
                        match = re.search(r'(\d+(\.\d+)?)', value)
                        if match:
                            row_data[field] = float(match.group(1))
                        else:
                            row_data[field] = value
                    elif value is not None:
                        try:
                            row_data[field] = float(value)
                        except (ValueError, TypeError):
                            row_data[field] = value
                    else:
                        row_data[field] = None

                elif field == 'volume':
                    if value and isinstance(value, str):
                        value = value.replace(',', '.')
                        try:
                            row_data[field] = float(value)
                        except (ValueError, TypeError):
                            row_data[field] = value
                    elif value is not None:
                        try:
                            row_data[field] = float(value)
                        except (ValueError, TypeError):
                            row_data[field] = value
                    else:
                        row_data[field] = None

                elif field == 'productName':
                    if not value:
                        row_data[field] = None
                        continue

                    # Всегда переводим с китайского на русский
                    if isinstance(value, str) and len(value.strip()) > 0:
                        try:
                            translated_value = GoogleTranslator(source='zh-TW', target='ru').translate(value)
                            row_data[field] = translated_value
                            logger.info(f"Переведено с китайского на русский: '{value}' -> '{translated_value}'")
                        except Exception as e:
                            logger.warning(f"Ошибка перевода с китайского: {e}")
                            row_data[field] = value
                    else:
                        row_data[field] = value

                elif field in ['freightTariff', 'insurance', 'packaging']:
                    if value and isinstance(value, str):
                        match = re.search(r'(\d+(\.\d+)?)', value)
                        if match:
                            row_data[field] = float(match.group(1))
                        else:
                            row_data[field] = value
                    elif value is not None:
                        try:
                            row_data[field] = float(value)
                        except (ValueError, TypeError):
                            row_data[field] = value
                    else:
                        row_data[field] = None

                elif field == 'total':
                    if value and isinstance(value, str):
                        match = re.search(r'(\d+(\.\d+)?)', value)
                        if match:
                            row_data[field] = float(match.group(1))
                        else:
                            # Если не удалось извлечь число, пробуем перевести и извлечь
                            try:
                                # Всегда переводим с китайского
                                translated_text = GoogleTranslator(source='zh-TW', target='ru').translate(value)
                                row_data[field] = translated_text
                            except Exception as e:
                                logger.warning(f"Ошибка перевода: {e}")
                                row_data[field] = value
                    elif value is not None:
                        try:
                            row_data[field] = float(value)
                        except (ValueError, TypeError):
                            row_data[field] = value
                    else:
                        row_data[field] = None

            # Проверяем, что у нас есть хотя бы код клиента
            if row_data.get('clientCode'):
                self.parsed_data.append(row_data)
            else:
                logger.warning(f"Пропущена строка {row}: отсутствует код клиента")

    def _calculate_batch_summary(self) -> Dict[str, Any]:
        """
        Рассчитывает сводную информацию по партии.
        Учитывает все значения в сборных местах при подсчете коробок, веса, объема и суммы.

        Returns:
            Dict[str, Any]: Словарь со сводной информацией
        """
        if not self.parsed_data:
            return {
                'total_places': 0,
                'total_composite_places': 0,
                'total_boxes': 0,
                'total_weight': 0,
                'total_volume': 0,
                'total_amount': 0,
                'unique_clients': 0
            }

        # Инициализируем счетчики
        total_places = 0
        total_composite_places = 0
        total_boxes = 0
        total_weight = 0
        total_volume = 0
        total_amount = 0

        # Множество для отслеживания уникальных числовых частей кодов клиентов
        unique_client_numeric_codes = set()

        # Множество для отслеживания уже учтенных групп сборных мест
        counted_composite_groups = set()

        # Обрабатываем каждую запись
        for item in self.parsed_data:
            client_code = item.get('clientCode')
            client_numeric_code = item.get('clientCodeNumeric')

            # Добавляем числовую часть кода клиента в множество уникальных клиентов
            if client_numeric_code:
                unique_client_numeric_codes.add(client_numeric_code)

            # Проверяем, является ли это сборным местом
            is_composite = item.get('isCompositeCargo', False)
            composite_group_id = item.get('compositeGroupId')

            # Для сборных мест считаем место только один раз
            if is_composite and composite_group_id:
                if composite_group_id not in counted_composite_groups:
                    total_places += 1
                    total_composite_places += 1
                    counted_composite_groups.add(composite_group_id)
            else:
                # Для обычных мест просто добавляем количество мест
                places_count = item.get('placesCount', 0) or 0
                total_places += places_count

            # Для всех остальных показателей учитываем значения из каждой строки
            # независимо от того, сборное это место или нет

            # Считаем коробки
            boxes_count = item.get('boxesCount', 0) or 0
            total_boxes += boxes_count

            # Считаем вес
            weight = item.get('weight', 0)
            if weight is not None and isinstance(weight, (int, float)):
                total_weight += weight

            # Считаем объем
            volume = item.get('volume', 0)
            if volume is not None and isinstance(volume, (int, float)):
                total_volume += volume

            # Считаем общую сумму
            amount = item.get('total', 0)
            if amount is not None and isinstance(amount, (int, float)):
                total_amount += amount

        # Формируем и возвращаем результат
        return {
            'total_places': total_places,
            'total_composite_places': total_composite_places,
            'total_boxes': total_boxes,
            'total_weight': round(total_weight, 2),
            'total_volume': round(total_volume, 3),
            'total_amount': round(total_amount, 2),
            'unique_clients': len(unique_client_numeric_codes)
        }

    def _print_results(self) -> None:
        """
        Выводит результаты парсинга в консоль в форматированном виде.
        Использует библиотеку tabulate для красивого вывода таблиц.
        """
        print("\n" + "=" * 80)
        print(f"РЕЗУЛЬТАТЫ ПАРСИНГА ФАЙЛА: {self.file_path}")
        print("=" * 80)

        print(f"\nНомер партии (баланса): {self.batch_number} (числовая часть: {self.batch_number_numeric})")
        print(f"Найдено записей: {len(self.parsed_data)}")

        # Выводим сводную информацию по партии
        print("\nСВОДНАЯ ИНФОРМАЦИЯ ПО ПАРТИИ:")
        print(f"Общее количество мест: {self.batch_summary.get('total_places', 0)}")
        print(f"Количество сборных мест: {self.batch_summary.get('total_composite_places', 0)}")
        print(f"Общее количество коробок: {self.batch_summary.get('total_boxes', 0)}")
        print(f"Общий вес: {self.batch_summary.get('total_weight', 0)} кг")
        print(f"Общий объём: {self.batch_summary.get('total_volume', 0)} м³")
        print(f"Общая сумма: {self.batch_summary.get('total_amount', 0)} $")
        print(f"Количество уникальных клиентов: {self.batch_summary.get('unique_clients', 0)}")

        if not self.parsed_data:
            print("\nНет данных для отображения.")
            return

        # Создаем таблицу для вывода основных данных
        table_data = []
        headers = ["№", "Клиент", "Мест", "Вес", "Объем", "Наименование", "Итого", "Сборное место ID"]

        for i, item in enumerate(self.parsed_data, 1):
            row = [
                i,
                item.get('clientCode', ''),
                item.get('placesCount', ''),
                item.get('weight', ''),
                item.get('volume', ''),
                (item.get('productName', '')[:30] + '...') if item.get('productName', '') and len(
                    item.get('productName', '')) > 30 else item.get('productName', ''),
                item.get('total', ''),
                item.get('compositeId', '')
            ]
            table_data.append(row)

        print("\nОбзор данных:")
        print(tabulate(table_data, headers=headers, tablefmt="grid"))

        # Выводим детали для первых 10 записей
        print("\nДетали первых записей:")
        for i, item in enumerate(self.parsed_data[:10], 1):
            print(f"\n--- Запись #{i} {'(Сборное место)' if item.get('isCompositeCargo') else ''} ---")

            # Выводим основные поля
            fields_to_print = [
                ('clientCode', 'ID клиента'),
                ('clientCodeNumeric', 'Числовая часть ID клиента'),
                ('depatureFromChinaDate', 'Дата выхода из Китая'),
                ('placesCount', 'Количество мест'),
                ('compositePlacesCount', 'Мест в сборном грузе' if item.get('isCompositeCargo') else None),
                ('compositeId', 'ID сборного места' if item.get('isCompositeCargo') else None),
                ('weight', 'Вес, кг'),
                ('boxesCount', 'Количество коробок'),
                ('cubicTariff', 'Тариф куб, $'),
                ('unitsCount', 'Кол-во ед., шт'),
                ('productPrice', 'Цена товара, ¥'),
                ('insurancePercent', 'Страховка, %'),
                ('volume', 'Объём, м3'),
                ('productName', 'Наименование'),
                ('freightTariff', 'Тариф фрахт, $'),
                ('insurance', 'Страховка, $'),
                ('packaging', 'Упаковка, $'),
                ('total', 'Итого, $')
            ]

            for field, label in fields_to_print:
                if label:  # Пропускаем None
                    value = item.get(field)
                    if value is not None:
                        print(f"{label}: {value}")

        print("\n" + "=" * 80)


def main(file_path: str) -> None:
    """
    Основная функция для запуска парсера из командной строки.

    Args:
        file_path: Путь к Excel файлу
    """
    if not os.path.exists(file_path):
        logger.error(f"Файл не найден: {file_path}")
        return

    parser = CargoExcelParser(file_path)
    parser.parse()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Использование: python CargoParserExcel.py <путь_к_excel_файлу>")
        sys.exit(1)

    file_path = sys.argv[1]
    main(file_path)
